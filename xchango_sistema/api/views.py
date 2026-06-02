from django.shortcuts import render
from rest_framework import viewsets
from django.http import HttpResponse
from rest_framework.decorators import action
from openpyxl import Workbook

from .models import (
    NivelAcceso,
    Permiso,
    Administrador,
    AdministradorPermiso,
    HistorialAdmin,
    Notificacion,
)

from .serializers import (
    NivelAccesoSerializer,
    PermisoSerializer,
    AdministradorSerializer,
    AdministradorPermisoSerializer,
    HistorialAdminSerializer,
    NotificacionSerializer,
)


# CRUD tabla nivel_acceso (GET, POST, PUT, DELETE)
class NivelAccesoViewSet(viewsets.ModelViewSet):
    queryset = NivelAcceso.objects.all()
    serializer_class = NivelAccesoSerializer

    @action(detail=False, methods=['get'])
    def excel(self, request):
        wb = Workbook()
        ws = wb.active

        ws.title = "Nivel Acceso"

        ws.append([
            "ID",
            "Nombre",
            "Activo",
            "Fecha Asignacion",
            "Fecha Creacion",
            "Fecha Modificacion"
        ])

        for item in self.get_queryset():
            ws.append([
                item.id_nivelacceso,
                item.nombre,
                item.activo,
                item.fecha_asignacion.strftime("%Y-%m-%d %H:%M:%S"),
                item.fecha_creacion.strftime("%Y-%m-%d %H:%M:%S"),
                item.fecha_modificacion.strftime("%Y-%m-%d %H:%M:%S")
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        response['Content-Disposition'] = (
            'attachment; filename="nivel_acceso.xlsx"'
        )

        wb.save(response)

        return response

# CRUD tabla permiso (GET, POST, PUT, DELETE)
class PermisoViewSet(viewsets.ModelViewSet):
    queryset = Permiso.objects.all()
    serializer_class = PermisoSerializer

    @action(detail=False, methods=['get'])
    def excel(self, request):
        wb = Workbook()
        ws = wb.active

        ws.title = "Permiso"
        ws.append([
            "ID",
            "Nombre",
            "Descripcion",
            "Activo",
            "Fecha Creacion",
            "Fecha Modificacion"
        ])

        for item in self.get_queryset():
            ws.append([
                item.id_permiso,
                item.nombre,
                item.descripcion,
                item.activo,
                item.fecha_creacion.strftime("%Y-%m-%d %H:%M:%S"),
                item.fecha_modificacion.strftime("%Y-%m-%d %H:%M:%S")
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        response['Content-Disposition'] = (
            'attachment; filename="permiso.xlsx"'
        )

        wb.save(response)

        return response


# CRUD tabla administrador (GET, POST, PUT, DELETE)
class AdministradorViewSet(viewsets.ModelViewSet):
    queryset = Administrador.objects.all()
    serializer_class = AdministradorSerializer

    @action(detail=False, methods=['get'])
    def excel(self, request):
        wb = Workbook()
        ws = wb.active

        ws.title = "Administrador"

        
        ws.append([
            "ID Admin",
            "ID Usuario",
            "Nivel Acceso",
            "Activo",
            "Fecha Asignacion",
            "Fecha Creacion",
            "Fecha Modificacion"
        ])

        
        for item in self.get_queryset():
            ws.append([
                item.id_admin,
                item.id_usuario,
                item.id_nivelacceso.nombre,
                item.activo,
                item.fecha_asignacion.strftime("%Y-%m-%d %H:%M:%S"),
                item.fecha_creacion.strftime("%Y-%m-%d %H:%M:%S"),
                item.fecha_modificacion.strftime("%Y-%m-%d %H:%M:%S")
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        response['Content-Disposition'] = (
            'attachment; filename="administrador.xlsx"'
        )

        wb.save(response)

        return response

# CRUD tabla administradorpermiso (GET, POST, PUT, DELETE)
class AdministradorPermisoViewSet(viewsets.ModelViewSet):
    queryset = AdministradorPermiso.objects.all()
    serializer_class = AdministradorPermisoSerializer

    @action(detail=False, methods=['get'])
    def excel(self, request):
        wb = Workbook()
        ws = wb.active

        ws.title = "Administrador Permiso"

        ws.append([
            "ID",
            "ID Admin",
            "ID Permiso",
            "Activo",
            "Fecha Creacion",
            "Fecha Modificacion"
        ])

        for item in self.get_queryset():
            ws.append([
                item.id,
                item.id_admin.id_admin,
                item.id_permiso.id_permiso,
                item.activo,
                item.fecha_creacion.strftime("%Y-%m-%d %H:%M:%S"),
                item.fecha_modificacion.strftime("%Y-%m-%d %H:%M:%S")
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        response['Content-Disposition'] = (
            'attachment; filename="administrador_permiso.xlsx"'
        )

        wb.save(response)

        return response
    



# CRUD tabla historialadmin (GET, POST, PUT, DELETE)
class HistorialAdminViewSet(viewsets.ModelViewSet):
    queryset = HistorialAdmin.objects.all()
    serializer_class = HistorialAdminSerializer

    @action(detail=False, methods=['get'])
    def excel(self, request):
        wb = Workbook()
        ws = wb.active

        ws.title = "Historial Admin"

        ws.append([
            "ID Historial",
            "ID Admin",
            "Accion",
            "Descripcion",
            "Tipo Objeto",
            "ID Objeto",
            "Fecha Accion",
            "Activo",
            "Fecha Creacion",
            "Fecha Modificacion"
        ])

        for item in self.get_queryset():
            ws.append([
                item.id_historial,
                item.id_admin.id_admin,
                item.accion,
                item.descripcion,
                item.tipo_objeto,
                item.id_objeto,
                item.fecha_accion.strftime("%Y-%m-%d %H:%M:%S"),
                item.activo,
                item.fecha_creacion.strftime("%Y-%m-%d %H:%M:%S"),
                item.fecha_modificacion.strftime("%Y-%m-%d %H:%M:%S")
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        response['Content-Disposition'] = (
            'attachment; filename="historial_admin.xlsx"'
        )

        wb.save(response)

        return response


# CRUD tabla notificacion (GET, POST, PUT, DELETE)
class NotificacionViewSet(viewsets.ModelViewSet):
    queryset = Notificacion.objects.all()
    serializer_class = NotificacionSerializer
    @action(detail=False, methods=['get'])
    def excel(self, request):
        wb = Workbook()
        ws = wb.active

        ws.title = "Notificaciones"

        ws.append([
            "ID Notificacion",
            "ID Usuario",
            "Titulo",
            "Mensaje",
            "Tipo",
            "ID Referencia",
            "Tipo Referencia",
            "Leido",
            "Activo",
            "Fecha Creacion",
            "Fecha Modificacion"
        ])

        for item in self.get_queryset():
            ws.append([
                item.id_notificacion,
                item.id_usuario,
                item.titulo,
                item.mensaje,
                item.tipo,
                item.id_referencia,
                item.tipo_referencia,
                item.leido,
                item.activo,
                item.fecha_creacion.strftime("%Y-%m-%d %H:%M:%S"),
                item.fecha_modificacion.strftime("%Y-%m-%d %H:%M:%S")
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        response['Content-Disposition'] = (
            'attachment; filename="notificaciones.xlsx"'
        )

        wb.save(response)

        return response

